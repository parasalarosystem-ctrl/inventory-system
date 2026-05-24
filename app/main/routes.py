from flask import Blueprint, render_template, request, redirect, url_for, flash, session, jsonify, current_app, send_file
from app.models import Inventory, DisposalHistory, Borrowing, User, AuditLog, POSTransaction, POSTransactionItem, Category
from app.extensions import db, login_manager, limiter
from flask_login import login_user, logout_user, login_required, current_user
import re
from datetime import datetime, timedelta, date
from io import BytesIO
from sqlalchemy import asc, desc, func
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
import os
from werkzeug.utils import secure_filename
from dateutil.relativedelta import relativedelta
from itsdangerous import URLSafeTimedSerializer
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

main = Blueprint('main', __name__, template_folder='templates')


@main.context_processor
def inject_notif_count():
    if not current_user.is_authenticated:
        return {'notif_count': 0, 'notif_ids': '[]'}
    now = date.today()
    low_items = Inventory.query.filter(
        Inventory.status == 'Active',
        Inventory.is_deleted == False,
        Inventory.quantity <= Inventory.restock_threshold
    ).with_entities(Inventory.id).all()
    exp_items = Inventory.query.filter(
        Inventory.status == 'Active',
        Inventory.is_deleted == False,
        Inventory.expiry_date != None,
        Inventory.expiry_date <= (now + timedelta(days=30))
    ).with_entities(Inventory.id).all()
    all_ids = list({r.id for r in low_items} | {r.id for r in exp_items})
    import json
    return {'notif_count': len(all_ids), 'notif_ids': json.dumps(all_ids)}


# ---------------- NOTIFICATIONS ----------------
@main.route('/notifications')
def notifications():
    now = date.today()

    from sqlalchemy import case as sa_case
    low_stock_items = Inventory.query.filter(
        Inventory.status == 'Active',
        Inventory.is_deleted == False,
        Inventory.quantity <= Inventory.restock_threshold
    ).order_by(
        sa_case(
            (Inventory.quantity <= 0, 1),
            else_=0
        ),
        Inventory.quantity.asc()
    ).all()

    expiration_items = []
    active_items = Inventory.query.filter_by(status='Active', is_deleted=False).all()
    for item in active_items:
        if item.expiry_date:
            delta = (item.expiry_date - now).days
            if delta < 0:
                expiration_items.append({'item': item, 'end_date': item.expiry_date, 'days_left': delta, 'status': 'Expired'})
            elif delta <= 30:
                expiration_items.append({'item': item, 'end_date': item.expiry_date, 'days_left': delta, 'status': 'Expiring Soon'})

    # Sort: expired first (most overdue), then expiring soonest
    expiration_items.sort(key=lambda x: (0 if x['status'] == 'Expired' else 1, x['days_left']))

    return render_template('notifications.html', low_stock_items=low_stock_items, expiration_items=expiration_items, today=now)

# ---------------- DASHBOARD ----------------
@main.route('/dashboard')
def dashboard():
    now = date.today()

    total_products = Inventory.query.filter(
        Inventory.status == 'Active',
        Inventory.is_deleted == False
    ).count()

    low_stock_count = Inventory.query.filter(
        Inventory.status == 'Active',
        Inventory.is_deleted == False,
        Inventory.quantity <= Inventory.restock_threshold
    ).count()

    disposed_count = Inventory.query.filter(
        Inventory.is_deleted == False,
        Inventory.quantity <= 0
    ).count()

    damaged_count = Inventory.query.filter_by(status='Damaged', is_deleted=False).count()
    lost_count = Inventory.query.filter_by(status='Lost', is_deleted=False).count()

    expiration_count = Inventory.query.filter(
        Inventory.status == 'Active',
        Inventory.is_deleted == False,
        Inventory.expiry_date != None,
        Inventory.expiry_date <= (now + timedelta(days=30))
    ).count()

    recent_records = DisposalHistory.query.order_by(DisposalHistory.event_date.desc()).limit(8).all()

    EVENT_MAP = {
        'Added':       ('bg-t-blue',   'fa-plus-circle',  'Product Added'),
        'Edited':      ('bg-t-orange', 'fa-edit',         'Product Edited'),
        'Disposal':    ('bg-t-red',    'fa-trash-alt',    'Product Disposed'),
        'Restoration': ('bg-t-green',  'fa-undo',         'Product Restored'),
        'Borrowed':    ('bg-t-blue',   'fa-hand-holding', 'Product Borrowed'),
        'Returned':    ('bg-t-green',  'fa-undo-alt',     'Product Returned'),
    }

    activities = []
    for r in recent_records:
        badge, icon, title = EVENT_MAP.get(r.event_type, ('bg-t-orange', 'fa-info-circle', r.event_type))
        try:
            formatted_date = r.event_date.strftime('%Y-%m-%d %H:%M') if r.event_date else '—'
        except Exception:
            formatted_date = str(r.event_date)
        activities.append({
            'date': formatted_date,
            'badge': badge,
            'icon': icon,
            'title': title,
            'description': f"{r.item.product_name} — {r.remarks or ''}" if r.item else r.remarks or '',
            'qty': int(r.quantity) if r.quantity else 1,
            'status': r.status_snapshot or '',
        })

    # ── Sales Prediction Data (last 6 months actuals + 3 months forecast) ──
    from calendar import month_abbr
    chart_labels = []
    chart_actuals = []

    # Build last 6 months of issued qty from DisposalHistory (Added events = stock in, issued = sales proxy)
    for i in range(5, -1, -1):
        m_date = now - relativedelta(months=i)
        label = f"{month_abbr[m_date.month]} {m_date.year}"
        chart_labels.append(label)

        # Use DisposalHistory Borrowed+Returned events as sales proxy per month
        month_sales = db.session.query(func.sum(DisposalHistory.quantity)).filter(
            DisposalHistory.event_type.in_(['Borrowed', 'Disposal', 'Deleted']),
            func.extract('year',  DisposalHistory.event_date) == m_date.year,
            func.extract('month', DisposalHistory.event_date) == m_date.month
        ).scalar() or 0

        chart_actuals.append(int(month_sales))

    # Simple linear regression forecast for next 3 months
    n = len(chart_actuals)
    x_vals = list(range(n))
    x_mean = sum(x_vals) / n
    y_mean = sum(chart_actuals) / n
    num = sum((x_vals[i] - x_mean) * (chart_actuals[i] - y_mean) for i in range(n))
    den = sum((x_vals[i] - x_mean) ** 2 for i in range(n))
    slope = num / den if den != 0 else 0
    intercept = y_mean - slope * x_mean

    chart_forecast = [None] * n  # null for past months
    for i in range(1, 4):
        m_date = now + relativedelta(months=i)
        chart_labels.append(f"{month_abbr[m_date.month]} {m_date.year}")
        predicted = max(0, round(intercept + slope * (n + i - 1)))
        chart_forecast.append(predicted)

    # Pad actuals with null for forecast months
    chart_actuals += [None, None, None]

    import json
    sales_chart = {
        'labels': json.dumps(chart_labels),
        'actuals': json.dumps(chart_actuals),
        'forecast': json.dumps(chart_forecast),
    }

    return render_template(
        'dashboard.html',
        total_items=total_products,
        low_stock=low_stock_count,
        disposed_count=disposed_count,
        damaged_count=damaged_count,
        lost_count=lost_count,
        expiration_count=expiration_count,
        recent_activity=activities,
        sales_chart=sales_chart,
    )


# ---------------- ROOT ROUTE ----------------
@main.route('/')
def home():
    if current_user.is_authenticated:
        return redirect(url_for('main.dashboard'))
    return redirect(url_for('main.login'))

# ---------------- INVENTORY LIST ----------------
@main.route('/inventory_list', methods=['GET'])
def inventory_list():

    query = Inventory.query.filter(
        Inventory.status == 'Active',
        Inventory.is_deleted == False
        
        ) 

    # --- Search ---
    search_term = request.args.get('search', '')
    if search_term:
        query = query.filter(
            Inventory.product_name.ilike(f'%{search_term}%')
        )
        flash(f'Showing results for "{search_term}".', 'info')

    # --- Sorting ---
    sort_by = request.args.get('sort_by', 'expiry_date')
    order = request.args.get('order', 'desc')

    if sort_by == 'expiry_date':
        query = query.order_by(asc(Inventory.expiry_date)) if order == 'asc' else query.order_by(desc(Inventory.expiry_date))
    elif sort_by == 'selling_price':
        query = query.order_by(asc(Inventory.selling_price)) if order == 'asc' else query.order_by(desc(Inventory.selling_price))
    elif sort_by == 'product_name':
        query = query.order_by(asc(Inventory.product_name)) if order == 'asc' else query.order_by(desc(Inventory.product_name))
    items = query.all()

    # Sort items by stock status: In Stock (qty > reorder) → Low Stock (0 < qty <= reorder) → Out of Stock (qty <= 0)
    def stock_sort_key(item):
        balance = item.quantity or 0
        reorder = item.restock_threshold or 5
        if balance > reorder:
            return 0  # In Stock - highest priority
        elif balance > 0:
            return 1  # Low Stock - medium priority
        else:
            return 2  # Out of Stock - lowest priority

    items = sorted(items, key=stock_sort_key)

    # Compute expiration status per item
    now = date.today()
    expiry_status = {}
    expiry_dates = {}
    for item in items:
        if item.expiry_date:
            try:
                delta = (item.expiry_date - now).days
                expiry_dates[item.id] = item.expiry_date.strftime('%Y-%m-%d')
                if delta < 0:
                    expiry_status[item.id] = 'expired'
                elif delta <= 30:
                    expiry_status[item.id] = 'expiring'
                else:
                    expiry_status[item.id] = 'good'
            except:
                expiry_status[item.id] = 'none'
                expiry_dates[item.id] = None
        else:
            expiry_status[item.id] = 'none'
            expiry_dates[item.id] = None

    return render_template(
        'index.html',
        items=items,
        expiry_status=expiry_status,
        expiry_dates=expiry_dates,
        highlight_id=request.args.get('highlight', ''),
        search_term=search_term,
        sort_by=sort_by,
        order=order
    )



@main.errorhandler(429)
def ratelimit_handler(e):
    flash("You have tried to log in too many times. Please wait 1 minute.", "error")
    return render_template("login.html")


# ---------------- LOGIN ----------------
@main.route('/login', methods=['GET', 'POST'])
@limiter.limit("5 per minute") 
def login():
    if current_user.is_authenticated:
        return redirect(url_for('main.dashboard'))

    if request.method == 'POST':
        identifier = request.form['username']
        password = request.form['password']

        user = User.query.filter(
            (User.username == identifier) | (User.email == identifier)
        ).first()

        if user and user.check_password(password):
            # Note: CAPTCHA is validated on client-side
            # The client-side JavaScript validates the CAPTCHA before form submission
            login_user(user)
            log_audit('Login', target=user.username or user.email, details='User logged in')
            flash(f'Welcome, {user.username or user.email}!', 'success')
            return redirect(url_for('main.dashboard'))
        else:
            flash('Invalid username/email or password', 'error')

    return render_template('login.html')

# ---------------- LOGOUT ----------------
@main.route('/logout')
def logout():
    logout_user()
    flash('You have been logged out.', 'info')
    return redirect(url_for('main.login'))


# ---------------- ADD PRODUCT ----------------
@main.route('/add_item', methods=['GET', 'POST'])
def add_item():
    if request.method == 'POST':
        product_name = request.form.get('product_name', '').strip()

        if not product_name:
            flash('Product Name is required.', 'error')
            return redirect(url_for('main.add_item'))

        date_val = request.form.get('expiry_date')
        if not date_val:
            flash('Expiry Date is required.', 'error')
            return redirect(url_for('main.add_item'))

        try:
            receipt_qty = int(request.form.get('quantity', 0))
            amount = float(request.form.get('selling_price', 0.0))
            reorder_point = int(request.form.get('restock_threshold', 5))
        except ValueError:
            flash('Invalid number entered.', 'error')
            return redirect(url_for('main.add_item'))

        item_date = datetime.strptime(date_val, '%Y-%m-%d').date()

        product_image = None
        image_file = request.files.get('image')
        if image_file and image_file.filename:
            import uuid
            filename = str(uuid.uuid4()) + '_' + secure_filename(image_file.filename)
            upload_path = os.path.join(current_app.root_path, 'static', 'uploads')
            os.makedirs(upload_path, exist_ok=True)
            image_file.save(os.path.join(upload_path, filename))
            product_image = filename

        new_item = Inventory(
            expiry_date=item_date,
            product_name=product_name,
            unit=request.form.get('unit', '').strip(),
            quantity=receipt_qty,
            selling_price=amount,
            restock_threshold=reorder_point,
            undertaking_property=request.form.get('undertaking_property', '').strip(),
            product_image=product_image,
            status='Active'
        )
        db.session.add(new_item)
        db.session.commit()

        # Log the Add action
        add_log = DisposalHistory(
            inventory_id=new_item.id,
            event_type='Added',
            quantity=receipt_qty,
            status_snapshot='Active',
            remarks=f'Product added: {item_description}',
            event_date=datetime.now(),
            is_resolved=True
        )
        db.session.add(add_log)
        db.session.commit()

        flash(f'Item "{product_name}" added successfully!', 'success')
        log_audit('Add Product', target=product_name, details=f'Qty: {receipt_qty}, Price: {amount}')

        return redirect(url_for('main.inventory_list'))

    return render_template('add_item.html')


# ---------------- EDIT ITEM ----------------
# ---------------- EDIT ITEM (MERGED & SECURED) ----------------

@main.route('/edit_item/<int:item_id>', methods=['POST'])
def edit_item(item_id):
    item = Inventory.query.get_or_404(item_id)

    date_val = request.form.get('expiry_date')
    item.expiry_date = datetime.strptime(date_val, '%Y-%m-%d').date() if date_val else None

    item.product_name = request.form.get('product_name', item.product_name)
    item.undertaking_property = request.form.get('undertaking_property', item.undertaking_property)
    item.unit = request.form.get('unit', item.unit)

    receipt_qty = int(request.form.get('quantity') or item.quantity or 0)
    amount = float(request.form.get('selling_price') or item.selling_price or 0)
    item.restock_threshold = int(request.form.get('restock_threshold') or item.restock_threshold or 5)

    item.quantity = receipt_qty
    item.selling_price = amount

    file = request.files.get('image')
    delete_image = request.form.get('delete_image')
    if delete_image == 'yes':
        item.product_image = None
    elif file and file.filename != '':
        filename = secure_filename(file.filename)
        upload_path = os.path.join(current_app.root_path, 'static', 'uploads')
        os.makedirs(upload_path, exist_ok=True)
        file.save(os.path.join(upload_path, filename))
        item.product_image = filename

    db.session.commit()

    edit_log = DisposalHistory(
        inventory_id=item.id,
        event_type='Edited',
        quantity=item.quantity,
        status_snapshot='Active',
        remarks=f'Product updated: {item.product_name}',
        event_date=datetime.now(),
        is_resolved=True
    )
    db.session.add(edit_log)

    db.session.commit()

    check_low_stock_flash(item)
    flash('Item updated successfully', 'success')
    log_audit('Edit Product', target=item.product_name, details=f'Qty: {item.quantity}, Price: {item.selling_price}')
    return redirect(url_for('main.inventory_list'))

# ---------------- DELETE PRODUCT ----------------
@main.route('/delete_item/<int:item_id>', methods=['POST'])
def delete_item(item_id):
    item = Inventory.query.get_or_404(item_id)
    item_name = item.product_name
    item_qty = item.quantity or 0

    item.is_deleted = True
    item.deleted_at = datetime.now()

    delete_log = DisposalHistory(
        inventory_id=item.id,
        event_type='Deleted',
        quantity=item_qty,
        status_snapshot='Deleted',
        remarks=f'Item deleted: {item_name}',
        event_date=datetime.now(),
        is_resolved=True
    )
    db.session.add(delete_log)

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting item: {str(e)}', 'error')
        return redirect(url_for('main.inventory_list'))

    log_audit('Delete Product', target=item_name, details=f'Qty: {item_qty} — Moved to Recycle Bin')
    flash('Item moved to Trash.', 'success')
    return redirect(url_for('main.inventory_list'))


# --- 3. NEW RESTORE FUNCTION ---
# --- UPDATED RESTORE FUNCTION (SMART MERGE) ---
@main.route('/restore_item/<int:item_id>', methods=['POST'])
def restore_item(item_id):
    # This is the item currently in the trash
    trash_item = Inventory.query.get_or_404(item_id)
    
    # Check if there is a matching ACTIVE item (same description)
    active_item = Inventory.query.filter(
        Inventory.product_name == trash_item.product_name,
        Inventory.is_deleted == False
    ).first()

    if active_item:
        # MERGE LOGIC: Add the trash quantity back to the existing active item
        active_item.quantity += trash_item.quantity

        # Permanently delete the "Trash Copy" since it's now merged
        db.session.delete(trash_item)
        flash(f'Restored {trash_item.quantity} items. Merged into existing inventory record.', 'success')

    else:
        # STANDARD LOGIC: No active match found, so just un-delete this row
        trash_item.is_deleted = False
        trash_item.deleted_at = None
        flash('Item restored as a separate entry in Inventory.', 'success')
    
    db.session.commit()
    log_audit('Restore Product', target=trash_item.product_name, details='Restored from Recycle Bin')
    return redirect(url_for('main.settings'))




# --- 4. NEW PERMANENT DELETE FUNCTION ---
@main.route('/force_delete_item/<int:item_id>', methods=['POST'])
def force_delete_item(item_id):
    item = Inventory.query.get_or_404(item_id)
    item_name = item.product_name
    db.session.delete(item) # Actually delete from DB
    db.session.commit()
    log_audit('Permanent Delete', target=item_name, details='Permanently removed from system')
    flash('Item permanently deleted.', 'success')
    return redirect(url_for('main.settings'))





# ---------------- GET ITEM DATA FOR PRINT ----------------
@main.route('/get_item_data_for_print/<int:item_id>', methods=['GET'])
def get_item_data_for_print(item_id):
    item = Inventory.query.get(item_id)
    if item:
        return jsonify({
            'id': item.id,
            'expiry_date': item.expiry_date,
            'product_name': item.product_name,
            'unit': item.unit,
            'quantity': item.quantity,
            'selling_price': item.selling_price,
            'undertaking_property': item.undertaking_property,
            'product_image': item.product_image
        })
    return jsonify({'error': 'Item not found'}), 404

# ---------------- VALIDATION UTILS ----------------
def validate_password(password):
    pattern = r'^(?=.*[a-z])(?=.*[A-Z])(?=.*\d)(?=.*[@$!%*?&])[A-Za-z\d@$!%*?&]{8,}$'
    return re.match(pattern, password)


def check_low_stock_flash(item):
    qty = item.quantity or 0
    reorder = item.restock_threshold or 5
    if qty <= 0:
        flash(f'Low Stock: "{item.product_name}" is Out of Stock.', 'warning')
    elif qty <= reorder:
        flash(f'Low Stock: "{item.product_name}" has only {qty} left.', 'warning')


def check_low_stock_flash(item):
    qty = item.quantity or 0
    reorder = item.restock_threshold or 5
    if qty <= 0:
        flash(f'Low Stock: "{item.product_name}" is Out of Stock.', 'warning')
    elif qty <= reorder:
        flash(f'Low Stock: "{item.product_name}" has only {qty} left.', 'warning')

# ---------------- AUDIT LOG HELPER ----------------
def log_audit(action, target=None, details=None):
    try:
        from flask import request as _req
        ip = _req.remote_addr
    except Exception:
        ip = None
    try:
        user = (current_user.username or current_user.email) if current_user.is_authenticated else 'System'
        entry = AuditLog(action=action, target=target, details=details, user=user, ip_address=ip)
        db.session.add(entry)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f'log_audit failed: {e}')


# ---------------- SALES CHART API (shared by Dashboard + Sales) ----------------
@main.route('/api/sales_chart')
def api_sales_chart():
    from calendar import month_abbr

    today = date.today()

    # ── Labels: last 3 months (actuals) + 1 month ahead (forecast) ──
    # Index 0-2 = actuals, index 3 = forecast
    monthly_labels = []
    for m in range(2, -1, -1):
        md = today - relativedelta(months=m)
        monthly_labels.append(f"{month_abbr[md.month]} {md.year}")
    forecast_month = today + relativedelta(months=1)
    monthly_labels.append(f"{month_abbr[forecast_month.month]} {forecast_month.year}")

    # Weekly labels kept for the Sales page toggle (last 8 weeks)
    weekly_labels = []
    for w in range(7, -1, -1):
        week_start = today - timedelta(days=today.weekday() + 7 * w)
        weekly_labels.append(week_start.strftime('%b %d'))

    palette = [
        '#3b82f6','#10b981','#f59e0b','#ef4444','#8b5cf6',
        '#06b6d4','#f97316','#84cc16','#ec4899','#14b8a6'
    ]

    # All event types that represent an inventory reduction (sale)
    SALE_EVENTS = ['Borrowed', 'Disposal', 'Deleted', 'Issued']

    sales_items = Inventory.query.filter(
        Inventory.status == 'Active',
        Inventory.is_deleted == False
    ).order_by(Inventory.product_name).all()

    # Sort by total items removed (from DisposalHistory) as a sales proxy
    def _total_sold(item):
        return db.session.query(func.sum(DisposalHistory.quantity)).filter(
            DisposalHistory.inventory_id == item.id,
            DisposalHistory.event_type.in_(['Borrowed', 'Disposal', 'Deleted', 'Issued'])
        ).scalar() or 0

    top_items = sorted(sales_items, key=_total_sold, reverse=True)[:10]

    products = []
    for idx, item in enumerate(top_items):

        # ── 3-month actuals ──
        monthly_actuals = []
        for m in range(2, -1, -1):
            md = today - relativedelta(months=m)
            qty = db.session.query(func.sum(DisposalHistory.quantity)).filter(
                DisposalHistory.inventory_id == item.id,
                DisposalHistory.event_type.in_(SALE_EVENTS),
                func.extract('year',  DisposalHistory.event_date) == md.year,
                func.extract('month', DisposalHistory.event_date) == md.month
            ).scalar() or 0
            monthly_actuals.append(int(qty))

        # ── 1-month forecast: average of the 3 actuals ──
        avg = round(sum(monthly_actuals) / len(monthly_actuals)) if any(monthly_actuals) else 0
        forecast_value = avg

        # ── Weekly actuals (for Sales page) ──
        weekly_data = []
        for w in range(7, -1, -1):
            week_start = today - timedelta(days=today.weekday() + 7 * w)
            week_end   = week_start + timedelta(days=6)
            qty = db.session.query(func.sum(DisposalHistory.quantity)).filter(
                DisposalHistory.inventory_id == item.id,
                DisposalHistory.event_type.in_(SALE_EVENTS),
                func.cast(DisposalHistory.event_date, db.Date) >= week_start,
                func.cast(DisposalHistory.event_date, db.Date) <= week_end
            ).scalar() or 0
            weekly_data.append(int(qty))

        products.append({
            'name':     item.product_name,
            'color':    palette[idx % len(palette)],
            # monthly: 3 actuals + None placeholder (forecast shown separately)
            'monthly':  monthly_actuals + [None],
            # forecast: None for the 3 actual months + projected value
            'forecast': [None, None, None, forecast_value],
            'weekly':   weekly_data,
            'unit':     item.unit or '',
            'price':    float(item.selling_price or 0),
            'balance':  int(item.quantity or 0),
        })

    return jsonify({
        'monthly_labels': monthly_labels,
        'weekly_labels':  weekly_labels,
        'products':       products,
    })
@main.route('/api/audit_logs')
def api_audit_logs():
    """Returns the latest audit log entries as JSON for real-time polling."""
    since_id = request.args.get('since_id', 0, type=int)
    limit = request.args.get('limit', 50, type=int)
    limit = min(limit, 2000)  # cap to prevent abuse

    query = AuditLog.query.order_by(desc(AuditLog.timestamp))
    if since_id:
        query = query.filter(AuditLog.id > since_id)
    logs = query.limit(limit).all()

    return jsonify([{
        'id': log.id,
        'timestamp': log.timestamp.strftime('%Y-%m-%d %H:%M:%S') if log.timestamp else '—',
        'user': log.user,
        'action': log.action,
        'target': log.target or '—',
        'details': log.details or '—',
        'ip_address': log.ip_address or '—'
    } for log in logs])


# ---------------- EXCEL EXPORT (Server Side) ----------------
@main.route('/export_excel')
def export_excel():
    items = Inventory.query.filter(Inventory.is_deleted == False).all()
    if not items:
        flash('No items to export.', 'warning')
        return redirect(url_for('main.inventory_list'))

    wb = Workbook()
    ws = wb.active
    ws.title = "PGPC Inventory"

    # Header Styling
    header_font = Font(name='Century Gothic', size=11, bold=True)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid") # Blue
    white_font = Font(name='Century Gothic', size=11, bold=True, color="FFFFFF")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    headers = [
        'EXPIRY DATE', 'PRODUCT NAME', 'CATEGORY', 'UNIT',
        'QUANTITY', 'SELLING PRICE', 'RESTOCK THRESHOLD', 'STATUS'
    ]
    ws.append(headers)

    col_widths = [12, 35, 20, 12, 12, 18, 18, 12]
    for col_num, cell in enumerate(ws[1]):
        cell.font = white_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align
        ws.column_dimensions[get_column_letter(col_num + 1)].width = col_widths[col_num]

    for item in items:
        row_data = [
            item.expiry_date,
            item.product_name,
            item.undertaking_property or '',
            item.unit or '',
            item.quantity or 0,
            item.selling_price or 0,
            item.restock_threshold or 5,
            item.status or 'Active',
        ]
        ws.append(row_data)

        for col_num, cell in enumerate(ws[ws.max_row], start=1):
            cell.border = thin_border
            cell.font = Font(name='Century Gothic', size=10)
            if col_num == 1:
                cell.number_format = 'YYYY-MM-DD'
                cell.alignment = center_align
            elif col_num == 6:
                cell.number_format = '₱#,##0.00'
                cell.alignment = Alignment(horizontal='right', vertical='center')
            elif col_num in [5, 7]:
                cell.number_format = '#,##0'
                cell.alignment = center_align
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # Add Filters
    ws.auto_filter.ref = ws.dimensions

    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    return send_file(
        excel_buffer,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='PGPC_Inventory_Export.xlsx'
    )

# ---------------- NEW IMPORT (VIA WEB/JS) ----------------
# ---------------- NEW IMPORT (VIA WEB/JS) ----------------


@main.route('/import_excel_data', methods=['POST'])
def import_excel_data():
    try:
        data = request.get_json()
        items = data.get('items', [])
        
        added_count = 0
        skipped_count = 0
        skipped_examples = []
        
        # To prevent duplicates within the Excel file itself
        seen_in_this_batch = set()

        for item in items:
            # 1. Clean and Prepare Data
            desc = str(item.get('product_name') or item.get('item_description', '')).strip()

            # Basic validation — product name is required
            if not desc:
                continue

            # -----------------------------------------------------------
            # CHECK IF ITEM ALREADY EXISTS (keyed by product name)
            # -----------------------------------------------------------
            exists_in_db = Inventory.query.filter(
                Inventory.product_name == desc,
                Inventory.is_deleted == False
            ).first()

            exists_in_batch = desc in seen_in_this_batch

            if exists_in_batch:
                skipped_count += 1
                if len(skipped_examples) < 3:
                    skipped_examples.append(desc)
                continue

            seen_in_this_batch.add(desc)

            # -----------------------------------------------------------
            # DATA PARSING
            # -----------------------------------------------------------
            try:
                incoming_receipt_qty = int(item.get('quantity') or item.get('receipt_qty', 1))
                incoming_amount = float(item.get('selling_price') or item.get('amount', 0))
                incoming_reorder = int(item.get('restock_threshold') or item.get('reorder_point', 5))
                incoming_undertaking = item.get('undertaking_property')
                incoming_life = item.get('unit') or item.get('estimated_useful_life')
                incoming_image = item.get('product_image') or item.get('image_filename') or None
            except ValueError:
                continue

            # Date parsing
            formatted_date = None
            raw_date = item.get('expiry_date') or item.get('date')
            if raw_date:
                try:
                    formatted_date = datetime.strptime(raw_date, '%Y-%m-%d').date()
                except:
                    formatted_date = None

            # -----------------------------------------------------------
            # MERGE OR CREATE ITEM
            # -----------------------------------------------------------
            if exists_in_db:
                # MERGE: Add incoming quantity to existing stock
                exists_in_db.quantity += incoming_receipt_qty
                added_count += 1
            else:
                new_item = Inventory(
                    expiry_date=formatted_date,
                    product_name=desc,
                    unit=incoming_life,
                    quantity=incoming_receipt_qty,
                    selling_price=incoming_amount,
                    undertaking_property=incoming_undertaking,
                    restock_threshold=incoming_reorder,
                    product_image=incoming_image,
                    status='Active'
                )
                db.session.add(new_item)
                added_count += 1
            
        db.session.commit()
        
        msg = f"Import Complete: {added_count} items added."
        if skipped_count > 0:
            msg += f" (Skipped {skipped_count} duplicates: {', '.join(skipped_examples)}...)"

        log_audit('Import Excel', target='Bulk Import', details=f'{added_count} items imported, {skipped_count} skipped')
        return jsonify({'status': 'success', 'message': msg})
        
    except Exception as e:
        db.session.rollback()
        print(f"Import Error: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500



# ---------------- REPORTS ROUTE ----------------



@main.route('/reports')
def reports():
    now = datetime.now().date()

    # ============================================================
    # 1. TRANSACTION HISTORY FOR REPORTS
    # ============================================================
    
    # Transaction History (For analytics and overview)
    transaction_logs = db.session.query(DisposalHistory, Inventory)\
        .join(Inventory, DisposalHistory.inventory_id == Inventory.id)\
        .order_by(desc(DisposalHistory.event_date)).limit(200).all()

    # Dedicated disposal/delete query — no limit, covers all records regardless of item status
    disposal_logs = db.session.query(DisposalHistory, Inventory)\
        .join(Inventory, DisposalHistory.inventory_id == Inventory.id)\
        .filter(DisposalHistory.event_type.in_(['Disposal', 'Deleted']))\
        .order_by(desc(DisposalHistory.event_date)).all()

    # ============================================================
    # 2. LOW STOCK REPORT (Kept the same)
    # ============================================================
    threshold = session.get('low_stock_threshold', 5) 
    
    # FIXED CODE — Out of Stock (balance <= 0) first, then Low Stock
    from sqlalchemy import case as sa_case
    low_stock_items = Inventory.query.filter(
        Inventory.status == 'Active',
        Inventory.quantity <= Inventory.restock_threshold
    ).order_by(
        sa_case(
            (Inventory.quantity <= 0, 1),
            else_=0
        ),
        Inventory.quantity.asc()
    ).all()

    # ============================================================
    # 3. EXPIRATION REPORT
    # ============================================================
    eol_items = []
    active_inventory = Inventory.query.filter(
        Inventory.status == 'Active',
        Inventory.is_deleted == False,
        Inventory.expiry_date != None
    ).all()

    for item in active_inventory:
        delta = (item.expiry_date - now).days
        if delta < 0:
            status_life = 'Expired'
            badge_color = 'bg-red'
        elif delta <= 30:
            status_life = f'Expiring Soon ({delta} days)'
            badge_color = 'bg-yellow'
        else:
            status_life = f'{delta} days left'
            badge_color = 'bg-green'

        eol_items.append({
            'category': item.undertaking_property or '—',
            'description': item.product_name,
            'useful_life': item.unit or '—',
            'quantity': item.quantity or 0,
            'end_date': item.expiry_date,
            'status_life': status_life,
            'badge_color': badge_color,
            'is_expired': delta < 0,
            'days_delta': delta
        })

    # Sort: expired first (most overdue at top), then closest expiry date
    eol_items.sort(key=lambda x: (0 if x['is_expired'] else 1, x['days_delta']))

    # ============================================================
    # 4. SALES DATA
    # ============================================================
    sales_items = Inventory.query.filter(
        Inventory.status == 'Active',
        Inventory.is_deleted == False
    ).order_by(Inventory.product_name).all()

    total_value = sum((i.quantity or 0) * (i.selling_price or 0) for i in sales_items)
    total_stock = sum((i.quantity or 0) for i in sales_items)
    total_sold  = db.session.query(func.sum(DisposalHistory.quantity)).filter(
        DisposalHistory.event_type.in_(['Borrowed', 'Disposal', 'Deleted', 'Issued'])
    ).scalar() or 0

    sales_summary = {
        'total_products': len(sales_items),
        'total_value': total_value,
        'total_stock': total_stock,
        'total_sold': int(total_sold),
    }

    # ============================================================
    # 5. SUMMARIES
    # ============================================================
    total_items_count = Inventory.query.count()
    total_value_asset = sum(
        (i.quantity or 0) * (i.selling_price or 0)
        for i in Inventory.query.filter_by(status='Active').all()
    )

    active_count = Inventory.query.filter_by(status='Active').count()
    lost_count = Inventory.query.filter_by(status='Lost').count()
    damaged_count = Inventory.query.filter_by(status='Damaged').count()
    disposed_count = Inventory.query.filter(Inventory.status.in_(['Disposed', 'Sold', 'Salvaged'])).count()

    financial_by_year = db.session.query(
        func.extract('year', Inventory.expiry_date).label('year'),
        func.sum(Inventory.quantity * Inventory.selling_price).label('total_value')
    ).filter(Inventory.status == 'Active').group_by('year').all()

    # ============================================================
    # 5. RENDER TEMPLATE (Updated Variables)
    # ============================================================
    audit_logs = AuditLog.query.order_by(desc(AuditLog.timestamp)).limit(200).all()

    return render_template(
        'reports.html',
        transaction_logs=transaction_logs,
        disposal_logs=disposal_logs,
        low_stock_items=low_stock_items,
        eol_items=eol_items,
        financial_by_year=financial_by_year,
        sales_items=sales_items,
        sales_summary=sales_summary,
        audit_logs=audit_logs,
        today=now,
        summary={
            'total_count': total_items_count,
            'total_value': total_value_asset,
            'active': active_count,
            'lost': lost_count,
            'damaged': damaged_count,
            'disposed': disposed_count
        }
    )



# ============================================================
# API: SALES PERFORMANCE DATA
# ============================================================
@main.route('/api/sales_performance')
def api_sales_performance():
    period = request.args.get('period', 'daily')
    start_date_str = request.args.get('start')
    end_date_str = request.args.get('end')
    
    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    except:
        return jsonify({'error': 'Invalid date format'}), 400
    
    # Sales events
    SALE_EVENTS = ['Borrowed', 'Disposal', 'Deleted', 'Issued']
    
    # Get sales logs for the period
    sales_logs = DisposalHistory.query.filter(
        DisposalHistory.event_type.in_(SALE_EVENTS),
        func.cast(DisposalHistory.event_date, db.Date) >= start_date,
        func.cast(DisposalHistory.event_date, db.Date) <= end_date
    ).order_by(desc(DisposalHistory.event_date)).all()
    
    # Calculate metrics
    total_qty = 0
    total_revenue = 0
    logs_data = []
    
    for log in sales_logs:
        qty = log.quantity or 0
        price = log.item.selling_price if log.item else 0
        value = qty * price

        total_qty += qty
        total_revenue += value

        logs_data.append({
            'date_time': log.event_date.strftime('%Y-%m-%d %H:%M') if log.event_date else '—',
            'product': log.item.product_name if log.item else '—',
            'category': log.item.undertaking_property if log.item else '—',
            'quantity': qty,
            'unit_price': price,
            'total_value': value
        })
    
    transactions = len(sales_logs)
    avg_revenue = total_revenue / transactions if transactions > 0 else 0
    
    return jsonify({
        'total_qty': total_qty,
        'total_revenue': total_revenue,
        'transactions': transactions,
        'avg_revenue': avg_revenue,
        'logs': logs_data
    })


# ---------------- SALES ROUTE ----------------
@main.route('/sales')
def sales():
    sales_items = Inventory.query.filter(
        Inventory.status == 'Active',
        Inventory.is_deleted == False
    ).order_by(Inventory.product_name).all()

    def stock_sort_key(item):
        qty = item.quantity or 0
        reorder = item.restock_threshold or 5
        if qty > reorder:
            return 0
        elif qty > 0:
            return 1
        else:
            return 2

    sales_items = sorted(sales_items, key=stock_sort_key)

    total_value = sum((i.quantity or 0) * (i.selling_price or 0) for i in sales_items)
    total_stock = sum((i.quantity or 0) for i in sales_items)
    total_sold  = db.session.query(func.sum(DisposalHistory.quantity)).filter(
        DisposalHistory.event_type.in_(['Borrowed', 'Disposal', 'Deleted', 'Issued'])
    ).scalar() or 0

    # ── Calculate Daily, Weekly, Monthly, Yearly Sales ──
    today = date.today()
    
    # All event types that represent an inventory reduction (sale)
    SALE_EVENTS = ['Borrowed', 'Disposal', 'Deleted', 'Issued']
    
    # ========== DAILY SALES ==========
    daily_sales_qty = db.session.query(func.sum(DisposalHistory.quantity)).filter(
        DisposalHistory.event_type.in_(SALE_EVENTS),
        func.cast(DisposalHistory.event_date, db.Date) == today
    ).scalar() or 0
    
    # Calculate daily sales value
    daily_sales_value = 0
    daily_logs = DisposalHistory.query.filter(
        DisposalHistory.event_type.in_(SALE_EVENTS),
        func.cast(DisposalHistory.event_date, db.Date) == today
    ).all()
    for log in daily_logs:
        if log.item:
            daily_sales_value += (log.quantity or 0) * (log.item.selling_price or 0)
    
    # Yesterday's sales for comparison
    yesterday = today - timedelta(days=1)
    yesterday_sales_qty = db.session.query(func.sum(DisposalHistory.quantity)).filter(
        DisposalHistory.event_type.in_(SALE_EVENTS),
        func.cast(DisposalHistory.event_date, db.Date) == yesterday
    ).scalar() or 0
    
    daily_change = 0
    if yesterday_sales_qty > 0:
        daily_change = ((daily_sales_qty - yesterday_sales_qty) / yesterday_sales_qty) * 100
    elif daily_sales_qty > 0:
        daily_change = 100
    
    # ========== WEEKLY SALES ==========
    week_start = today - timedelta(days=6)
    weekly_sales_qty = db.session.query(func.sum(DisposalHistory.quantity)).filter(
        DisposalHistory.event_type.in_(SALE_EVENTS),
        func.cast(DisposalHistory.event_date, db.Date) >= week_start,
        func.cast(DisposalHistory.event_date, db.Date) <= today
    ).scalar() or 0
    
    # Calculate weekly sales value
    weekly_sales_value = 0
    weekly_logs = DisposalHistory.query.filter(
        DisposalHistory.event_type.in_(SALE_EVENTS),
        func.cast(DisposalHistory.event_date, db.Date) >= week_start,
        func.cast(DisposalHistory.event_date, db.Date) <= today
    ).all()
    for log in weekly_logs:
        if log.item:
            weekly_sales_value += (log.quantity or 0) * (log.item.selling_price or 0)
    
    # Previous week for comparison
    prev_week_start = today - timedelta(days=13)
    prev_week_end = today - timedelta(days=7)
    prev_week_sales_qty = db.session.query(func.sum(DisposalHistory.quantity)).filter(
        DisposalHistory.event_type.in_(SALE_EVENTS),
        func.cast(DisposalHistory.event_date, db.Date) >= prev_week_start,
        func.cast(DisposalHistory.event_date, db.Date) <= prev_week_end
    ).scalar() or 0
    
    weekly_change = 0
    if prev_week_sales_qty > 0:
        weekly_change = ((weekly_sales_qty - prev_week_sales_qty) / prev_week_sales_qty) * 100
    elif weekly_sales_qty > 0:
        weekly_change = 100
    
    # ========== MONTHLY SALES ==========
    monthly_sales_qty = db.session.query(func.sum(DisposalHistory.quantity)).filter(
        DisposalHistory.event_type.in_(SALE_EVENTS),
        func.extract('year', DisposalHistory.event_date) == today.year,
        func.extract('month', DisposalHistory.event_date) == today.month
    ).scalar() or 0
    
    # Calculate monthly sales value
    monthly_sales_value = 0
    monthly_logs = DisposalHistory.query.filter(
        DisposalHistory.event_type.in_(SALE_EVENTS),
        func.extract('year', DisposalHistory.event_date) == today.year,
        func.extract('month', DisposalHistory.event_date) == today.month
    ).all()
    for log in monthly_logs:
        if log.item:
            monthly_sales_value += (log.quantity or 0) * (log.item.selling_price or 0)
    
    # Previous month for comparison
    prev_month = today - relativedelta(months=1)
    prev_month_sales_qty = db.session.query(func.sum(DisposalHistory.quantity)).filter(
        DisposalHistory.event_type.in_(SALE_EVENTS),
        func.extract('year', DisposalHistory.event_date) == prev_month.year,
        func.extract('month', DisposalHistory.event_date) == prev_month.month
    ).scalar() or 0
    
    monthly_change = 0
    if prev_month_sales_qty > 0:
        monthly_change = ((monthly_sales_qty - prev_month_sales_qty) / prev_month_sales_qty) * 100
    elif monthly_sales_qty > 0:
        monthly_change = 100
    
    # ========== YEARLY SALES ==========
    yearly_sales_qty = db.session.query(func.sum(DisposalHistory.quantity)).filter(
        DisposalHistory.event_type.in_(SALE_EVENTS),
        func.extract('year', DisposalHistory.event_date) == today.year
    ).scalar() or 0
    
    # Calculate yearly sales value
    yearly_sales_value = 0
    yearly_logs = DisposalHistory.query.filter(
        DisposalHistory.event_type.in_(SALE_EVENTS),
        func.extract('year', DisposalHistory.event_date) == today.year
    ).all()
    for log in yearly_logs:
        if log.item:
            yearly_sales_value += (log.quantity or 0) * (log.item.selling_price or 0)
    
    # Previous year for comparison
    prev_year_sales_qty = db.session.query(func.sum(DisposalHistory.quantity)).filter(
        DisposalHistory.event_type.in_(SALE_EVENTS),
        func.extract('year', DisposalHistory.event_date) == today.year - 1
    ).scalar() or 0
    
    yearly_change = 0
    if prev_year_sales_qty > 0:
        yearly_change = ((yearly_sales_qty - prev_year_sales_qty) / prev_year_sales_qty) * 100
    elif yearly_sales_qty > 0:
        yearly_change = 100
    
    # ========== TRANSACTION COUNTS ==========
    daily_transactions = len(daily_logs)
    weekly_transactions = len(weekly_logs)
    monthly_transactions = len(monthly_logs)
    yearly_transactions = len(yearly_logs)
    
    # ========== AVERAGE TRANSACTION VALUES ==========
    daily_avg = daily_sales_value / daily_transactions if daily_transactions > 0 else 0
    weekly_avg = weekly_sales_value / weekly_transactions if weekly_transactions > 0 else 0
    monthly_avg = monthly_sales_value / monthly_transactions if monthly_transactions > 0 else 0
    yearly_avg = yearly_sales_value / yearly_transactions if yearly_transactions > 0 else 0

    sales_summary = {
        'total_products': len(sales_items),
        'total_value': total_value,
        'total_stock': total_stock,
        'total_sold': int(total_sold),
        # Daily
        'daily_sales_qty': int(daily_sales_qty),
        'daily_sales_value': daily_sales_value,
        'daily_change': round(daily_change, 1),
        'daily_transactions': daily_transactions,
        'daily_avg': daily_avg,
        # Weekly
        'weekly_sales_qty': int(weekly_sales_qty),
        'weekly_sales_value': weekly_sales_value,
        'weekly_change': round(weekly_change, 1),
        'weekly_transactions': weekly_transactions,
        'weekly_avg': weekly_avg,
        # Monthly
        'monthly_sales_qty': int(monthly_sales_qty),
        'monthly_sales_value': monthly_sales_value,
        'monthly_change': round(monthly_change, 1),
        'monthly_transactions': monthly_transactions,
        'monthly_avg': monthly_avg,
        # Yearly
        'yearly_sales_qty': int(yearly_sales_qty),
        'yearly_sales_value': yearly_sales_value,
        'yearly_change': round(yearly_change, 1),
        'yearly_transactions': yearly_transactions,
        'yearly_avg': yearly_avg,
    }

    # ── Per-product time-series for the interactive chart ──
    import json
    from calendar import month_abbr

    # Weekly: last 8 weeks (Mon–Sun buckets)
    weekly_labels = []
    for w in range(7, -1, -1):
        week_start = today - timedelta(days=today.weekday() + 7 * w)
        weekly_labels.append(week_start.strftime('%b %d'))

    # Monthly: last 6 months
    monthly_labels = []
    for m in range(5, -1, -1):
        md = today - relativedelta(months=m)
        monthly_labels.append(f"{month_abbr[md.month]} {md.year}")

    chart_products = []
    # Limit to top 10 by total sold to keep chart readable
    top_items = sorted(sales_items, key=lambda i: db.session.query(func.sum(DisposalHistory.quantity)).filter(
        DisposalHistory.inventory_id == i.id,
        DisposalHistory.event_type.in_(['Borrowed', 'Disposal', 'Deleted', 'Issued'])
    ).scalar() or 0, reverse=True)[:10]

    # Palette for lines
    palette = [
        '#3b82f6','#10b981','#f59e0b','#ef4444','#8b5cf6',
        '#06b6d4','#f97316','#84cc16','#ec4899','#14b8a6'
    ]

    for idx, item in enumerate(top_items):
        weekly_data = []
        for w in range(7, -1, -1):
            week_start = today - timedelta(days=today.weekday() + 7 * w)
            week_end   = week_start + timedelta(days=6)
            qty = db.session.query(func.sum(DisposalHistory.quantity)).filter(
                DisposalHistory.inventory_id == item.id,
                DisposalHistory.event_type.in_(SALE_EVENTS),
                func.cast(DisposalHistory.event_date, db.Date) >= week_start,
                func.cast(DisposalHistory.event_date, db.Date) <= week_end
            ).scalar() or 0
            weekly_data.append(int(qty))

        monthly_data = []
        for m in range(5, -1, -1):
            md = today - relativedelta(months=m)
            qty = db.session.query(func.sum(DisposalHistory.quantity)).filter(
                DisposalHistory.inventory_id == item.id,
                DisposalHistory.event_type.in_(SALE_EVENTS),
                func.extract('year',  DisposalHistory.event_date) == md.year,
                func.extract('month', DisposalHistory.event_date) == md.month
            ).scalar() or 0
            monthly_data.append(int(qty))

        chart_products.append({
            'name': item.product_name,
            'color': palette[idx % len(palette)],
            'weekly': weekly_data,
            'monthly': monthly_data,
            'unit': item.unit or '',
            'price': float(item.selling_price or 0),
        })

    sales_chart = {
        'weekly_labels':  json.dumps(weekly_labels),
        'monthly_labels': json.dumps(monthly_labels),
        'products':       json.dumps(chart_products),
    }

    return render_template('sales.html', sales_items=sales_items, sales_summary=sales_summary, sales_chart=sales_chart)


# ---------------- RESOLVE DISPOSAL ROUTE ----------------
# ---------------- RESOLVE DISPOSAL ROUTE ----------------


@main.route('/resolve_disposal', methods=['POST'])
def resolve_disposal():
    history_id = request.form.get('history_id')
    resolution_type = request.form.get('resolution_type')
    notes = request.form.get('notes')

    log = DisposalHistory.query.get_or_404(history_id)
    item = Inventory.query.get_or_404(log.inventory_id)

    if log.is_resolved:
        return redirect(url_for('main.reports'))

    qty_to_restore = log.quantity

    # 1. RESTORE LOGIC: Add quantity back to stock
    item.quantity += qty_to_restore

    # 2. Reset Status
    item.status = 'Active'

    # 3. Update Log
    log.is_resolved = True
    log.remarks += f" [Resolved: {resolution_type}]"

    # 4. Create Restoration Log
    restore_log = DisposalHistory(
        inventory_id=item.id,
        event_type='Restoration',
        quantity=qty_to_restore,
        status_snapshot='Active',
        remarks=f"Restored: {notes}",
        event_date=datetime.now(),
        is_resolved=True
    )
    
    db.session.add(restore_log)
    db.session.commit()

    log_audit('Resolve Disposal', target=item.product_name, details=f'Resolution: {resolution_type}, Notes: {notes}')
    flash('Item restored to Inventory.', 'success')
    return redirect(url_for('main.reports'))


# ---------------- DISPOSE ITEM ROUTE ----------------
# ---------------- DISPOSE ITEM ROUTE (UPDATED) ----------------


@main.route('/dispose_item', methods=['POST'])
def dispose_item():
    item_id = request.form.get('item_id')
    source = request.form.get('disposal_source') # 'balance' or 'issued'
    status = request.form.get('status')
    reason = request.form.get('reason')
    
    # In Asset-Based, quantity is usually just 1
    qty = int(request.form.get('dispose_qty', 1))

    item = Inventory.query.get_or_404(item_id)

    # Deduct from current stock
    if item.quantity < qty:
        flash('Error: Not enough stock to dispose.', 'error')
        return redirect(url_for('main.inventory_list'))

    item.quantity -= qty
    item.status = status 

    # 3. Create History Log
    new_log = DisposalHistory(
        inventory_id=item.id,
        event_type='Disposal',
        quantity=qty,
        status_snapshot=status,
        remarks=f"{reason} (Source: {source})",
        event_date=datetime.now(),
        is_resolved=False
    )
    
    db.session.add(new_log)
    db.session.commit()
    
    check_low_stock_flash(item)
    flash(f'Item marked as {status}.', 'success')
    log_audit('Dispose Product', target=item.product_name, details=f'Status: {status}, Qty: {qty}, Reason: {reason}')
    return redirect(url_for('main.inventory_list'))



@main.route('/fix_missing_history')
def fix_missing_history():
    # Find items that are NOT Active but have NO history
    items = Inventory.query.filter(Inventory.status != 'Active').all()
    count = 0
    
    for item in items:
        # Check if history exists
        history = DisposalHistory.query.filter_by(inventory_id=item.id).first()
        if not history:
            # Create the missing history record
            missing_log = DisposalHistory(
                inventory_id=item.id,
                event_type='Disposal',
                status_snapshot=item.status,
                event_date=datetime.now().date(),
                remarks="Legacy Disposal Data"
            )
            db.session.add(missing_log)
            count += 1
    
    db.session.commit()
    return f"Fixed {count} missing history records. Go to Reports now."


@main.route('/delete_multiple_items', methods=['POST'])
def delete_multiple_items():
    # Get the list of IDs from the form
    item_ids = request.form.getlist('item_ids')
    
    if not item_ids:
        flash('No items selected for deletion.', 'warning')
        return redirect(url_for('main.inventory_list'))

    try:
        deleted_count = 0
        for i_id in item_ids:
            item = Inventory.query.get(i_id)
            if item:
                # --- CHANGE STARTS HERE ---
                
                # OLD WAY (Permanent Delete):
                # db.session.delete(item) 

                # NEW WAY (Soft Delete / Move to Trash):
                item.is_deleted = True
                item.deleted_at = datetime.now()
                delete_log = DisposalHistory(
                    inventory_id=item.id,
                    event_type='Deleted',
                    quantity=item.quantity or 0,
                    status_snapshot='Deleted',
                    remarks=f'Item deleted: {item.product_name}',
                    event_date=datetime.now(),
                    is_resolved=True
                )
                db.session.add(delete_log)
                deleted_count += 1
        
        db.session.commit()
        flash(f'Successfully moved {deleted_count} items to the Recycle Bin.', 'success')
        log_audit('Bulk Delete', target='Multiple Items', details=f'{deleted_count} items moved to Recycle Bin')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting items: {str(e)}', 'error')

    return redirect(url_for('main.inventory_list'))


# ---------------- BACKUP & RECOVERY ----------------
@main.route('/backup_data')
def backup_data():
    inventory_items = Inventory.query.all()
    history_logs = DisposalHistory.query.all()
    borrowings = Borrowing.query.all()

    def serialize_date(val):
        if val is None:
            return None
        if hasattr(val, 'isoformat'):
            return val.isoformat()
        return str(val)

    backup = {
        'backup_date': datetime.now().isoformat(),
        'inventory': [{
            'id': i.id, 'expiry_date': serialize_date(i.expiry_date),
            'product_name': i.product_name, 'unit': i.unit,
            'quantity': i.quantity, 'selling_price': i.selling_price,
            'restock_threshold': i.restock_threshold,
            'undertaking_property': i.undertaking_property,
            'product_image': i.product_image, 'status': i.status,
            'is_deleted': i.is_deleted, 'deleted_at': serialize_date(i.deleted_at)
        } for i in inventory_items],
        'disposal_history': [{
            'id': h.id, 'inventory_id': h.inventory_id, 'event_type': h.event_type,
            'status_snapshot': h.status_snapshot, 'event_date': serialize_date(h.event_date),
            'remarks': h.remarks, 'quantity': h.quantity, 'is_resolved': h.is_resolved
        } for h in history_logs],
        'borrowings': [{
            'id': b.id, 'inventory_id': b.inventory_id, 'borrower_name': b.borrower_name,
            'quantity': b.quantity, 'status': b.status,
            'date_borrowed': serialize_date(b.date_borrowed),
            'date_returned': serialize_date(b.date_returned)
        } for b in borrowings]
    }

    import json
    buf = BytesIO(json.dumps(backup, indent=2).encode('utf-8'))
    buf.seek(0)
    filename = f"inventory_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    return send_file(buf, mimetype='application/json', as_attachment=True, download_name=filename)


@main.route('/restore_data', methods=['POST'])
def restore_data():
    import json
    file = request.files.get('backup_file')
    if not file:
        flash('No backup file provided.', 'error')
        return redirect(url_for('main.settings'))

    try:
        data = json.loads(file.read().decode('utf-8'))

        inv_count = hist_count = borrow_count = 0

        # Restore Inventory
        existing_inv_ids = {i.id for i in Inventory.query.with_entities(Inventory.id).all()}
        for item in data.get('inventory', []):
            if item['id'] in existing_inv_ids:
                continue
            new = Inventory(
                id=item['id'],
                expiry_date=datetime.strptime(item['expiry_date'], '%Y-%m-%d').date() if item.get('expiry_date') else (datetime.strptime(item['date'], '%Y-%m-%d').date() if item.get('date') else None),
                product_name=item.get('product_name') or item.get('item_description'),
                unit=item.get('unit') or item.get('estimated_useful_life'),
                quantity=item.get('quantity') or item.get('receipt_qty', 0),
                selling_price=item.get('selling_price') or item.get('amount', 0),
                restock_threshold=item.get('restock_threshold') or item.get('reorder_point', 5),
                undertaking_property=item.get('undertaking_property'),
                product_image=item.get('product_image') or item.get('image_filename'),
                status=item.get('status', 'Active'),
                is_deleted=item.get('is_deleted', False),
                deleted_at=datetime.fromisoformat(item['deleted_at']) if item.get('deleted_at') else None
            )
            db.session.add(new)
            inv_count += 1
        db.session.flush()

        # Restore Disposal History
        existing_hist_ids = {h.id for h in DisposalHistory.query.with_entities(DisposalHistory.id).all()}
        for h in data.get('disposal_history', []):
            if h['id'] in existing_hist_ids:
                continue
            db.session.add(DisposalHistory(
                id=h['id'], inventory_id=h['inventory_id'], event_type=h.get('event_type'),
                status_snapshot=h.get('status_snapshot'),
                event_date=datetime.fromisoformat(h['event_date']) if h.get('event_date') else None,
                remarks=h.get('remarks'), quantity=h.get('quantity', 1),
                is_resolved=h.get('is_resolved', False)
            ))
            hist_count += 1

        # Restore Borrowings
        existing_borrow_ids = {b.id for b in Borrowing.query.with_entities(Borrowing.id).all()}
        for b in data.get('borrowings', []):
            if b['id'] in existing_borrow_ids:
                continue
            db.session.add(Borrowing(
                id=b['id'], inventory_id=b['inventory_id'], borrower_name=b.get('borrower_name'),
                quantity=b.get('quantity', 1), status=b.get('status', 'Borrowed'),
                date_borrowed=datetime.fromisoformat(b['date_borrowed']) if b.get('date_borrowed') else None,
                date_returned=datetime.fromisoformat(b['date_returned']) if b.get('date_returned') else None
            ))
            borrow_count += 1

        db.session.commit()
        flash(f'Restore complete: {inv_count} inventory items, {hist_count} history logs, {borrow_count} borrowing records restored.', 'success')

    except Exception as e:
        db.session.rollback()
        flash(f'Restore failed: {str(e)}', 'error')

    return redirect(url_for('main.settings'))


# ---------------- SETTINGS PAGE ----------------
@main.route('/settings')
def settings():
    # 1. Fetch current threshold (existing logic)
    current_threshold = session.get('low_stock_threshold', 5) 

    # 2. AUTOMATIC CLEANUP: Check for items deleted > 30 days ago
    thirty_days_ago = datetime.now() - timedelta(days=30)
    
    # Find old items
    expired_items = Inventory.query.filter(
        Inventory.is_deleted == True, 
        Inventory.deleted_at < thirty_days_ago
    ).all()
    
    # Permanently delete them
    if expired_items:
        for item in expired_items:
            db.session.delete(item)
        db.session.commit()

    # 3. Fetch remaining trash items to display
    deleted_items = Inventory.query.filter_by(is_deleted=True).order_by(Inventory.deleted_at.desc()).all()
    
    # 4. Pass variables to template
    return render_template(
        'settings.html', 
        threshold=current_threshold, 
        deleted_items=deleted_items
    )

# ---------------- UPDATE PROFILE ----------------
# ---------------- UPDATE PROFILE (Simplified - No Image) ----------------
@main.route('/settings/update_profile', methods=['POST'])
def update_profile():
    new_username = request.form.get('username')
    new_email = request.form.get('email') # <--- Get the email from the form

    # 1. Validation: Check if username is taken
    existing_user = User.query.filter(User.username == new_username).first()
    if existing_user and existing_user.id != current_user.id:
        flash('Username already taken.', 'error')
        return redirect(url_for('main.settings'))

    # 2. Validation: Check if Email is taken (Optional, but good practice)
    if new_email:
        existing_email = User.query.filter(User.email == new_email).first()
        if existing_email and existing_email.id != current_user.id:
            flash('Email already registered by another user.', 'error')
            return redirect(url_for('main.settings'))

    # 3. SAVE TO DATABASE
    current_user.username = new_username
    current_user.email = new_email  # <--- This saves the Gmail to the DB
    
    db.session.commit()
    
    flash('Profile and Recovery Email updated successfully.', 'success')
    return redirect(url_for('main.settings'))

# ---------------- FORGOT PASSWORD (DEBUG MODE) ----------------




@main.route('/forgot_password', methods=['GET', 'POST'])
def forgot_password():
    if request.method == 'POST':
        email = request.form.get('email')
        
        # 1. Search for the user
        user = User.query.filter_by(email=email).first()

        if user:
            # --- EMAIL CONFIGURATION ---
            SYSTEM_EMAIL = "pgpcsystem010@gmail.com"
            SYSTEM_PASSWORD = "anfnvutiehfcvzic" 
            # ---------------------------

            # 2. Generate the secure link
            s = URLSafeTimedSerializer(current_app.config['SECRET_KEY'])
            token = s.dumps(user.email, salt='password-reset-salt')
            reset_link = url_for('main.reset_password', token=token, _external=True)

            # 3. Prepare the Email
            msg = MIMEMultipart()
            msg['From'] = SYSTEM_EMAIL
            msg['To'] = email
            msg['Subject'] = "Reset Your Password - Inventory System"

            body = f"""
            <div style="font-family: Arial, sans-serif; padding: 20px;">
                <h2>Password Reset Request</h2>
                <p>Hello,</p>
                <p>Click the button below to choose a new password:</p>
                <a href="{reset_link}">Reset Password</a>
                <p>If you did not request this, please ignore this email.</p>
            </div>
            """
            msg.attach(MIMEText(body, 'html'))

            # 4. Send the Email
            try:
                server = smtplib.SMTP('smtp.gmail.com', 587)
                server.starttls()
                server.login(SYSTEM_EMAIL, SYSTEM_PASSWORD)
                server.send_message(msg)
                server.quit()
                flash(f'A password reset link has been sent to {email}. Please check your inbox.', 'success')
            except Exception as e:
                print(f"Email Error: {e}")
                flash('Could not send email. Please check your internet connection and try again.', 'error')
                
            return redirect(url_for('main.forgot_password'))
        else:
            # Don't reveal whether email exists — security best practice
            flash('If that email is registered, a reset link has been sent.', 'success')
            return redirect(url_for('main.forgot_password'))
            
    return render_template('forgot_password.html')




@main.route('/reset_password/<token>', methods=['GET', 'POST'])
def reset_password(token):
    s = URLSafeTimedSerializer(current_app.config['SECRET_KEY'])
    try:
        email = s.loads(token, salt='password-reset-salt', max_age=3600)
    except:
        flash('Link expired or invalid.', 'error')
        return redirect(url_for('main.forgot_password'))

    if request.method == 'POST':
        password = request.form.get('password')
        confirm = request.form.get('confirm_password')
        if password != confirm:
            flash('Passwords do not match.', 'error')
            return render_template('reset_password.html', token=token)
        
        user = User.query.filter_by(email=email).first()
        if user:
            user.set_password(password)
            db.session.commit()
            flash('Password reset successful!', 'success')
            return redirect(url_for('main.login'))

    return render_template('reset_password.html', token=token)










# ---------------- CHANGE PASSWORD ----------------
@main.route('/settings/change_password', methods=['POST'])
def change_password():
    current_password = request.form.get('current_password')
    new_password = request.form.get('new_password')
    confirm_password = request.form.get('confirm_password')

    # Verify current password
    if not current_user.check_password(current_password):
        flash('Incorrect current password.', 'error')
        return redirect(url_for('main.settings'))

    if new_password != confirm_password:
        flash('New passwords do not match.', 'error')
        return redirect(url_for('main.settings'))
        
    if len(new_password) < 8:
        flash('Password must be at least 8 characters.', 'error')
        return redirect(url_for('main.settings'))

    # Set new password
    current_user.set_password(new_password)
    db.session.commit()
    
    flash('Password changed successfully!', 'success')
    return redirect(url_for('main.settings'))

# ---------------- SYSTEM PREFERENCES ----------------


# ---------------- NEW: PARTIAL DELETE ROUTE ----------------
@main.route('/delete_item_quantity', methods=['POST'])
def delete_item_quantity():
    item_id = request.form.get('item_id')
    delete_all = request.form.get('delete_all')
    
    try:
        qty_to_delete = int(request.form.get('quantity'))
    except (ValueError, TypeError):
        qty_to_delete = 0

    item = Inventory.query.get_or_404(item_id)

    # CASE A: Delete the Entire Record
    if delete_all == 'yes' or (qty_to_delete >= item.quantity):
        item.is_deleted = True
        item.deleted_at = datetime.now()

        delete_log = DisposalHistory(
            inventory_id=item.id,
            event_type='Deleted',
            quantity=item.quantity or 0,
            status_snapshot='Deleted',
            remarks=f'Entire item deleted: {item.product_name}',
            event_date=datetime.now(),
            is_resolved=True
        )
        db.session.add(delete_log)
        db.session.commit()

        log_audit('Delete Product', target=item.product_name,
                  details=f'Entire record deleted — Qty: {item.quantity or 0}')
        flash('Entire item record moved to Recycle Bin.', 'success')

    # CASE B: Partial Delete (Split the item)
    else:
        # 1. Create a "Ghost Copy" that represents the deleted items
        trash_copy = Inventory(
            expiry_date=item.expiry_date,
            product_name=item.product_name,
            unit=item.unit,
            quantity=qty_to_delete,
            selling_price=item.selling_price,
            undertaking_property=item.undertaking_property,
            product_image=item.product_image,
            is_deleted=True,
            deleted_at=datetime.now()
        )
        db.session.add(trash_copy)
        db.session.flush()

        # 2. Update the Original Active Item
        item.quantity = item.quantity - qty_to_delete

        # Log the partial deletion against the trash copy
        delete_log = DisposalHistory(
            inventory_id=trash_copy.id,
            event_type='Deleted',
            quantity=qty_to_delete,
            status_snapshot='Deleted',
            remarks=f'Partial delete: {item.product_name} (-{qty_to_delete})',
            event_date=datetime.now(),
            is_resolved=True
        )
        db.session.add(delete_log)
        db.session.commit()

        log_audit('Delete Product', target=item.product_name,
                  details=f'Partial delete — Qty removed: {qty_to_delete}')
        flash(f'Moved {qty_to_delete} items to Recycle Bin. Remaining stock updated.', 'success')

    return redirect(url_for('main.inventory_list'))


@main.route('/handle_bulk_trash', methods=['POST'])
def handle_bulk_trash():
    action = request.form.get('action') # 'restore' or 'delete'
    item_ids = request.form.getlist('trash_ids') # List of selected IDs

    if not item_ids:
        flash('No items selected.', 'warning')
        return redirect(url_for('main.settings'))

    try:
        if action == 'restore':
            count = 0
            for i_id in item_ids:
                trash_item = Inventory.query.get(i_id)
                if trash_item:
                    # Check for active duplicate (Same logic as single restore)
                    active_item = Inventory.query.filter(
                        Inventory.product_name == trash_item.product_name,
                        Inventory.is_deleted == False
                    ).first()

                    if active_item:
                        # Merge
                        active_item.quantity += trash_item.quantity
                        db.session.delete(trash_item)
                    else:
                        # Restore
                        trash_item.is_deleted = False
                        trash_item.deleted_at = None
                    count += 1
            flash(f'Successfully restored {count} items.', 'success')

        elif action == 'delete':
            count = 0
            for i_id in item_ids:
                item = Inventory.query.get(i_id)
                if item:
                    # Delete related borrowing records first
                    Borrowing.query.filter_by(inventory_id=item.id).delete()
                    # Delete related disposal history records
                    DisposalHistory.query.filter_by(inventory_id=item.id).delete()
                    # Now delete the item
                    db.session.delete(item)
                    count += 1
            flash(f'Permanently deleted {count} items.', 'success')

        db.session.commit()

    except Exception as e:
        db.session.rollback()
        flash(f'Error processing bulk action: {str(e)}', 'error')

    return redirect(url_for('main.settings'))


# ---------------- BORROW ITEM ROUTE ----------------
# ---------------- BORROW ITEM ROUTE ----------------
@main.route('/borrow_item', methods=['POST'])
def borrow_item():
    item_id = request.form.get('item_id')
    borrower_name = request.form.get('borrower_name')
    borrow_qty_raw = request.form.get('borrow_qty')

    # 1. Validation: Did we get all the data?
    if not item_id or not borrower_name or not borrow_qty_raw:
        flash('Error: Missing information. Please fill out all fields.', 'error')
        return redirect(url_for('main.inventory_list'))

    try:
        borrow_qty = int(borrow_qty_raw)
    except ValueError:
        flash('Error: Quantity must be a valid number.', 'error')
        return redirect(url_for('main.inventory_list'))
    
    # 2. Find the Item
    item = Inventory.query.get_or_404(item_id)
    
    # 3. Check Stock
    if (item.quantity or 0) < borrow_qty:
        flash(f'Error: Not enough stock! Current stock is {item.quantity or 0}.', 'error')
        return redirect(url_for('main.inventory_list'))

    # 4. Deduct stock
    item.quantity -= borrow_qty

    # 5. Create Borrowing Record
    new_loan = Borrowing(
        inventory_id=item.id,
        borrower_name=borrower_name,
        quantity=borrow_qty,
        status='Borrowed',
        date_borrowed=datetime.now()
    )
    
    # 6. Create Transaction Log
    borrow_log = DisposalHistory(
        inventory_id=item.id,
        event_type='Borrowed',
        quantity=borrow_qty,
        status_snapshot='Borrowed',
        remarks=f'Borrowed by {borrower_name}',
        event_date=datetime.now(),
        is_resolved=False
    )
    
    db.session.add(new_loan)
    db.session.add(borrow_log)
    db.session.commit()
    
    check_low_stock_flash(item)
    flash(f'{borrow_qty} item(s) borrowed by {borrower_name}.', 'success')
    log_audit('Borrow Product', target=item.product_name, details=f'Borrower: {borrower_name}, Qty: {borrow_qty}')
    return redirect(url_for('main.inventory_list'))
# ---------------- RETURN ITEM ROUTE ----------------
# ---------------- RETURN ITEM ROUTE ----------------
@main.route('/return_item/<int:borrow_id>', methods=['POST'])
def return_item(borrow_id):
    # Find the specific transaction
    loan_record = Borrowing.query.get_or_404(borrow_id)
    
    if loan_record.status == 'Returned':
        flash('Error: This item was already returned.', 'error')
        return redirect(url_for('main.borrowed_items'))

    # Find the original item in Inventory
    item = Inventory.query.get(loan_record.inventory_id)
    
    # 1. Update Inventory Master (Return to stock)
    item.quantity += loan_record.quantity

    # 2. Update Transaction Record
    loan_record.status = 'Returned'
    loan_record.date_returned = datetime.now()
    
    # 3. Create Return Log
    return_log = DisposalHistory(
        inventory_id=item.id,
        event_type='Returned',
        quantity=loan_record.quantity,
        status_snapshot='Active',
        remarks=f'Returned by {loan_record.borrower_name}',
        event_date=datetime.now(),
        is_resolved=True
    )
    
    db.session.add(return_log)
    db.session.commit()
    log_audit('Return Product', target=item.product_name, details=f'Returned by {loan_record.borrower_name}, Qty: {loan_record.quantity}')
    flash('Item returned to inventory successfully.', 'success')
    return redirect(url_for('main.borrowed_items'))

# ---------------- BORROWED ITEMS PAGE ----------------
@main.route('/borrowed_items')
@main.route('/borrow_list')
def borrowed_items():
    # Fetch borrowing records with inventory details
    borrowed_records = db.session.query(Borrowing, Inventory)\
        .join(Inventory, Borrowing.inventory_id == Inventory.id)\
        .filter(Borrowing.status == 'Borrowed')\
        .order_by(desc(Borrowing.date_borrowed)).all()
    return render_template('borrowed.html', borrowed_records=borrowed_records)

# ---------------- DISPOSAL & ISSUES PAGE ----------------
@main.route('/disposal_issues')
def disposal_issues():
    # Get disposal logs that need resolution
    disposal_logs = db.session.query(DisposalHistory, Inventory)\
        .join(Inventory, DisposalHistory.inventory_id == Inventory.id)\
        .filter(DisposalHistory.event_type.in_(['Disposal']))\
        .filter(DisposalHistory.is_resolved == False)\
        .order_by(desc(DisposalHistory.event_date)).all()
    
    # Get all transaction history for reference
    all_transactions = db.session.query(DisposalHistory, Inventory)\
        .join(Inventory, DisposalHistory.inventory_id == Inventory.id)\
        .order_by(desc(DisposalHistory.event_date)).all()
    
    return render_template('disposal_issues.html', 
                         disposal_logs=disposal_logs,
                         all_transactions=all_transactions)


# ════════════════════════════════════════════════════════════════
#  POS TERMINAL
# ════════════════════════════════════════════════════════════════

@main.route('/pos')
def pos_terminal():
    """Main POS terminal page — serves all active, in-stock products."""
    products = Inventory.query.filter(
        Inventory.status == 'Active',
        Inventory.is_deleted == False,
        Inventory.quantity > 0
    ).order_by(Inventory.product_name).all()

    # Build unique category list from undertaking_property field
    raw_cats = [p.undertaking_property for p in products if p.undertaking_property]
    categories = sorted(set(raw_cats))

    return render_template('pos_terminal.html', products=products, categories=categories)


@main.route('/pos/products')
def pos_products_api():
    """JSON endpoint — returns products filtered by category / search."""
    category = request.args.get('category', 'all')
    search   = request.args.get('search', '').strip()

    query = Inventory.query.filter(
        Inventory.status == 'Active',
        Inventory.is_deleted == False,
        Inventory.quantity > 0
    )

    if category and category != 'all':
        query = query.filter(Inventory.undertaking_property == category)

    if search:
        query = query.filter(Inventory.product_name.ilike(f'%{search}%'))

    items = query.order_by(Inventory.product_name).all()

    return jsonify([{
        'id':          i.id,
        'name':        i.product_name,
        'price':       float(i.selling_price or 0),
        'balance':     int(i.quantity or 0),
        'reorder':     int(i.restock_threshold or 5),
        'category':    i.undertaking_property or 'Uncategorized',
        'unit':        i.unit or '',
        'image':       i.product_image or '',
    } for i in items])


@main.route('/pos/complete_sale', methods=['POST'])
def pos_complete_sale():
    """Process a completed POS sale: deduct stock, save transaction, return receipt data."""
    data = request.get_json(silent=True)
    if not data or not data.get('items'):
        return jsonify({'success': False, 'error': 'Empty cart'}), 400

    cart_items      = data['items']          # [{id, name, price, qty}, ...]
    amount_tendered = float(data.get('amount_tendered', 0))
    total_amount    = float(data.get('total', 0))

    if amount_tendered < total_amount:
        return jsonify({'success': False, 'error': 'Insufficient payment'}), 400

    # Generate receipt number
    receipt_no = f"RCP-{datetime.now().strftime('%Y%m%d%H%M%S')}-{current_user.id}"

    try:
        txn = POSTransaction(
            receipt_no      = receipt_no,
            cashier         = current_user.username or current_user.email,
            total_amount    = total_amount,
            amount_tendered = amount_tendered,
            change_given    = round(amount_tendered - total_amount, 2),
            created_at      = datetime.now()
        )
        db.session.add(txn)
        db.session.flush()   # get txn.id before committing

        receipt_lines = []

        for ci in cart_items:
            inv_id  = int(ci['id'])
            qty     = int(ci['qty'])
            price   = float(ci['price'])
            name    = ci['name']
            sub     = round(price * qty, 2)

            item = Inventory.query.get(inv_id)
            if not item:
                db.session.rollback()
                return jsonify({'success': False, 'error': f'Product "{name}" not found'}), 404

            if (item.quantity or 0) < qty:
                db.session.rollback()
                return jsonify({'success': False,
                                'error': f'Not enough stock for "{name}". Available: {item.quantity or 0}'}), 400

            # Deduct stock
            item.quantity -= qty

            # History log
            db.session.add(DisposalHistory(
                inventory_id    = item.id,
                event_type      = 'Issued',
                quantity        = qty,
                status_snapshot = 'Active',
                remarks         = f'POS Sale — Receipt {receipt_no}',
                event_date      = datetime.now(),
                is_resolved     = True
            ))

            # Transaction line item
            db.session.add(POSTransactionItem(
                transaction_id = txn.id,
                inventory_id   = item.id,
                product_name   = name,
                unit_price     = price,
                quantity       = qty,
                subtotal       = sub
            ))

            receipt_lines.append({
                'name':     name,
                'qty':      qty,
                'price':    price,
                'subtotal': sub,
            })

        db.session.commit()

        log_audit('POS Sale', target=receipt_no,
                  details=f'Total: ₱{total_amount:.2f}, Items: {len(cart_items)}')

        return jsonify({
            'success':          True,
            'receipt_no':       receipt_no,
            'cashier':          txn.cashier,
            'created_at':       txn.created_at.strftime('%b %d, %Y %I:%M %p'),
            'items':            receipt_lines,
            'total':            total_amount,
            'amount_tendered':  amount_tendered,
            'change':           txn.change_given,
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@main.route('/pos/transactions')
def pos_transactions():
    """Transaction history page."""
    page     = request.args.get('page', 1, type=int)
    per_page = 20
    txns     = POSTransaction.query.order_by(
                   desc(POSTransaction.created_at)
               ).paginate(page=page, per_page=per_page, error_out=False)

    return render_template('pos_transactions.html', txns=txns)


@main.route('/pos/receipt/<int:txn_id>')
def pos_receipt(txn_id):
    """View a saved receipt."""
    txn = POSTransaction.query.get_or_404(txn_id)
    return render_template('pos_receipt.html', txn=txn)
